from __future__ import annotations

import unittest
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - dependency should be installed by project setup
    Workbook = None

from form_builder.excel_import_models import ImportJob
from form_builder.excel_import_services import (
    analyze_workbook,
    build_default_mapping,
    build_hygiene_demo_workbook_bytes,
    generate_draft_forms_from_mapping,
    safe_unique_key,
)
from form_builder.models import Form
from form_builder.repeatable_models import RepeatableGroup


@unittest.skipUnless(Workbook is not None, "openpyxl is required for Excel import tests")
@override_settings(DEFAULT_FILE_STORAGE="django.core.files.storage.FileSystemStorage")
class ExcelImportAnalyzerTests(TestCase):
    def workbook_bytes(self, builder) -> bytes:
        workbook = Workbook()
        builder(workbook)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_single_sheet_analysis_detects_title_header_and_inputs(self):
        def build(workbook):
            ws = workbook.active
            ws.title = "Hygiene"
            ws.merge_cells("A1:D1")
            ws["A1"] = "Hygiene Kontrolle"
            ws["A3"] = "Bereich"
            ws["B3"] = "OK"
            ws["C3"] = "Nicht OK"
            ws["D3"] = "Bemerkung"
            ws["A4"] = "Kueche"
            ws["B4"] = "☐"
            ws["C4"] = "☐"

        analysis = analyze_workbook(
            BytesIO(self.workbook_bytes(build)), original_filename="hygiene.xlsx"
        )
        sheet = analysis["sheets"][0]
        self.assertEqual(sheet["name"], "Hygiene")
        self.assertTrue(sheet["title_cells"])
        self.assertTrue(sheet["header_rows"])
        self.assertTrue(sheet["tables"])
        self.assertIn("merged_cells", sheet)

    def test_multiple_sheets_and_empty_workbook_are_deterministic(self):
        def build(workbook):
            workbook.active.title = "Leer"
            workbook.create_sheet("Daten")["A1"] = "Titel"

        analysis = analyze_workbook(
            BytesIO(self.workbook_bytes(build)), original_filename="multi.xlsx"
        )
        self.assertEqual(analysis["sheet_count"], 2)
        empty_sheet = analysis["sheets"][0]
        self.assertEqual(empty_sheet["row_count"], 0)
        self.assertEqual(empty_sheet["used_range"], "")

    def test_safe_unique_key_handles_duplicate_labels(self):
        used = set()
        first = safe_unique_key("Name", used)
        used.add(first)
        second = safe_unique_key("Name", used)
        self.assertEqual(first, "name")
        self.assertEqual(second, "name-2")


@unittest.skipUnless(Workbook is not None, "openpyxl is required for Excel import tests")
@override_settings(DEFAULT_FILE_STORAGE="django.core.files.storage.FileSystemStorage")
class ExcelImportGenerationTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("excel-admin", password="x", is_staff=True)

    def create_job_from_bytes(self, payload: bytes, name: str = "hygiene.xlsx") -> ImportJob:
        uploaded = SimpleUploadedFile(
            name,
            payload,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        job = ImportJob.objects.create(
            uploaded_file=uploaded,
            original_filename=name,
            uploaded_by=self.user,
            status=ImportJob.ImportStatus.ANALYZED,
        )
        job.uploaded_file.open("rb")
        job.analysis_result = analyze_workbook(job.uploaded_file, original_filename=name)
        job.uploaded_file.close()
        job.mapping = build_default_mapping(job.analysis_result, mode="all_sheets_one_form")
        job.save(update_fields=["analysis_result", "mapping", "updated_at"])
        return job

    def test_mapping_generates_draft_form_with_repeatable_table(self):
        job = self.create_job_from_bytes(build_hygiene_demo_workbook_bytes())
        forms = generate_draft_forms_from_mapping(job=job, user=self.user)
        self.assertEqual(len(forms), 1)
        form = forms[0]
        self.assertEqual(form.status, Form.PublicationStatus.DRAFT)
        self.assertTrue(form.sections.exists())
        self.assertTrue(RepeatableGroup.objects.filter(form=form).exists())

    def test_100_sheets_basic_success(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        for index in range(1, 101):
            ws = workbook.create_sheet(f"Sheet {index}")
            ws["A1"] = f"Titel {index}"
            ws["A3"] = "Feld"
            ws["B3"] = "OK"
            ws["A4"] = "Wert"
            ws["B4"] = "x"
        output = BytesIO()
        workbook.save(output)
        analysis = analyze_workbook(BytesIO(output.getvalue()), original_filename="many.xlsx")
        self.assertEqual(analysis["sheet_count"], 100)

    def test_hygiene_generation_adds_required_if_metadata(self):
        job = self.create_job_from_bytes(build_hygiene_demo_workbook_bytes())
        form = generate_draft_forms_from_mapping(job=job, user=self.user)[0]
        required_if_columns = []
        for group in RepeatableGroup.objects.filter(form=form):
            required_if_columns.extend(
                column
                for column in group.columns.all()
                if (column.validation_rules or {}).get("required_if")
            )
        self.assertTrue(required_if_columns)
