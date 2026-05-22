from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from time import monotonic
from typing import Any

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify

from .excel_import_models import ImportJob, ImportedSheet, SUPPORTED_EXCEL_EXTENSIONS
from .models import AuditLog, Field, Form, FormSection
from .repeatable_models import RepeatableGroup, RepeatableGroupColumn

MAX_EXCEL_IMPORT_SIZE = 12 * 1024 * 1024
MAX_SHEETS = 100
MAX_ROWS_PER_SHEET = 500
MAX_COLUMNS_PER_SHEET = 60
MAX_CELLS_TOTAL = 60000
MAX_ANALYSIS_SECONDS = 15
HEADER_SCORE_MIN = 2

CHECKBOX_VALUES = {
    "x",
    "yes",
    "no",
    "ja",
    "nein",
    "ok",
    "nicht ok",
    "true",
    "false",
    "wahr",
    "falsch",
    "☑",
    "☒",
    "☐",
}
DATE_HINTS = {"datum", "frist", "geburt", "date", "termin", "faellig", "fällig"}
NUMBER_HINTS = {"nr", "nummer", "anzahl", "menge", "betrag", "summe", "preis"}
LONG_TEXT_HINTS = {"bemerkung", "beschreibung", "mangel", "maßnahme", "massnahme", "notiz", "hinweis"}
PHOTO_HINTS = {"foto", "bild", "datei", "anlage", "upload"}
SIGNATURE_HINTS = {"unterschrift", "signatur"}


@dataclass(frozen=True)
class SheetBounds:
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def row_count(self) -> int:
        return max(self.max_row - self.min_row + 1, 0)

    @property
    def column_count(self) -> int:
        return max(self.max_col - self.min_col + 1, 0)


def ensure_openpyxl():
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on deployment packaging
        raise ValidationError(
            "openpyxl ist nicht installiert. Bitte Dependency installieren und erneut versuchen."
        ) from exc
    return load_workbook, get_column_letter


def validate_excel_upload(uploaded_file) -> None:
    name = getattr(uploaded_file, "name", "") or ""
    ext = os.path.splitext(name.lower())[1]
    if ext not in SUPPORTED_EXCEL_EXTENSIONS:
        raise ValidationError("Nur .xlsx-Dateien ohne Makros werden akzeptiert.")
    size = int(getattr(uploaded_file, "size", 0) or 0)
    if size > MAX_EXCEL_IMPORT_SIZE:
        raise ValidationError("Die Excel-Datei ist zu gross fuer den sicheren Import.")


def create_import_job_from_upload(*, uploaded_file, user) -> ImportJob:
    validate_excel_upload(uploaded_file)
    job = ImportJob.objects.create(
        uploaded_file=uploaded_file,
        original_filename=getattr(uploaded_file, "name", "workbook.xlsx")[:255],
        uploaded_by=user,
        status=ImportJob.ImportStatus.UPLOADED,
    )
    analyze_import_job(job)
    return job


def analyze_import_job(job: ImportJob) -> ImportJob:
    try:
        job.uploaded_file.open("rb")
        analysis = analyze_workbook(job.uploaded_file, original_filename=job.original_filename)
    except Exception as exc:
        job.status = ImportJob.ImportStatus.FAILED
        job.error_message = str(exc)
        job.save(update_fields=["status", "error_message", "updated_at"])
        raise
    finally:
        try:
            job.uploaded_file.close()
        except Exception:
            pass

    with transaction.atomic():
        job.analysis_result = analysis
        job.mapping = build_default_mapping(analysis)
        job.status = ImportJob.ImportStatus.ANALYZED
        job.error_message = ""
        job.save(update_fields=["analysis_result", "mapping", "status", "error_message", "updated_at"])
        ImportedSheet.objects.filter(job=job).delete()
        for index, sheet in enumerate(analysis.get("sheets", []), start=1):
            ImportedSheet.objects.create(
                job=job,
                sheet_index=index,
                name=sheet.get("name", f"Sheet {index}"),
                used_range=sheet.get("used_range") or "",
                row_count=sheet.get("row_count") or 0,
                column_count=sheet.get("column_count") or 0,
                analysis=sheet,
                selected=True,
            )
    return job


def analyze_workbook(file_obj, *, original_filename: str = "workbook.xlsx") -> dict:
    load_workbook, _get_column_letter = ensure_openpyxl()
    started = monotonic()
    workbook = load_workbook(file_obj, read_only=False, data_only=True, keep_links=False)
    sheet_names = workbook.sheetnames[:MAX_SHEETS]
    total_cells = 0
    sheets = []
    for sheet_index, sheet_name in enumerate(sheet_names, start=1):
        if monotonic() - started > MAX_ANALYSIS_SECONDS:
            raise ValidationError("Excel-Analyse wurde aus Sicherheitsgruenden abgebrochen.")
        worksheet = workbook[sheet_name]
        bounds = _used_bounds(worksheet)
        total_cells += min(bounds.row_count, MAX_ROWS_PER_SHEET) * min(
            bounds.column_count, MAX_COLUMNS_PER_SHEET
        )
        if total_cells > MAX_CELLS_TOTAL:
            raise ValidationError("Excel-Datei enthaelt zu viele Zellen fuer einen sicheren Import.")
        sheets.append(_analyze_sheet(worksheet, sheet_index=sheet_index, bounds=bounds))
    return {
        "version": 1,
        "source_filename": original_filename,
        "analyzed_at": timezone.now().isoformat(),
        "limits": {
            "max_sheets": MAX_SHEETS,
            "max_rows_per_sheet": MAX_ROWS_PER_SHEET,
            "max_columns_per_sheet": MAX_COLUMNS_PER_SHEET,
            "max_cells_total": MAX_CELLS_TOTAL,
        },
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def _cell_value(cell) -> Any:
    value = cell.value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _as_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _used_bounds(worksheet) -> SheetBounds:
    if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet.cell(1, 1).value is None:
        return SheetBounds(1, 0, 1, 0)
    max_row = min(worksheet.max_row, MAX_ROWS_PER_SHEET)
    max_col = min(worksheet.max_column, MAX_COLUMNS_PER_SHEET)
    return SheetBounds(worksheet.min_row, max_row, worksheet.min_column, max_col)


def _coordinate(cell) -> str:
    return cell.coordinate


def _used_range(worksheet, bounds: SheetBounds) -> str:
    if bounds.row_count == 0 or bounds.column_count == 0:
        return ""
    return f"{worksheet.cell(bounds.min_row, bounds.min_col).coordinate}:{worksheet.cell(bounds.max_row, bounds.max_col).coordinate}"


def _analyze_sheet(worksheet, *, sheet_index: int, bounds: SheetBounds) -> dict:
    merged_cells = [str(rng) for rng in worksheet.merged_cells.ranges]
    title_cells = _detect_title_cells(worksheet, bounds, merged_cells)
    header_rows = _detect_header_rows(worksheet, bounds)
    tables = _detect_tables(worksheet, bounds, header_rows)
    checkbox_like_cells = _detect_checkbox_like_cells(worksheet, bounds)
    date_like_cells = _detect_date_like_cells(worksheet, bounds)
    empty_input_cells = _detect_empty_input_cells(worksheet, bounds, tables)
    detected_fields = _detected_fields_from_inputs(empty_input_cells, date_like_cells, checkbox_like_cells)
    return {
        "index": sheet_index,
        "name": worksheet.title,
        "used_range": _used_range(worksheet, bounds),
        "row_count": bounds.row_count,
        "column_count": bounds.column_count,
        "merged_cells": merged_cells,
        "title_cells": title_cells,
        "header_rows": header_rows,
        "tables": tables,
        "checkbox_like_cells": checkbox_like_cells[:150],
        "date_like_cells": date_like_cells[:150],
        "empty_input_cells": empty_input_cells[:250],
        "detected_fields": detected_fields[:250],
    }


def _detect_title_cells(worksheet, bounds: SheetBounds, merged_cells: list[str]) -> list[dict]:
    titles = []
    max_scan_row = min(bounds.max_row, bounds.min_row + 7)
    for row in range(bounds.min_row, max_scan_row + 1):
        for col in range(bounds.min_col, bounds.max_col + 1):
            cell = worksheet.cell(row, col)
            value = _as_text(cell.value)
            if not value or len(value) > 160:
                continue
            is_merged = any(cell.coordinate in rng for rng in merged_cells)
            has_title_style = bool(getattr(cell.font, "bold", False) or is_merged or row <= 2)
            if has_title_style and any(ch.isalpha() for ch in value):
                titles.append({"cell": cell.coordinate, "value": value, "merged": is_merged})
    return titles[:8]


def _row_values(worksheet, row: int, bounds: SheetBounds) -> list[str]:
    return [_as_text(worksheet.cell(row, col).value) for col in range(bounds.min_col, bounds.max_col + 1)]


def _detect_header_rows(worksheet, bounds: SheetBounds) -> list[dict]:
    headers = []
    max_scan_row = min(bounds.max_row, bounds.min_row + 40)
    for row in range(bounds.min_row, max_scan_row + 1):
        values = _row_values(worksheet, row, bounds)
        non_empty = [value for value in values if value]
        if len(non_empty) < HEADER_SCORE_MIN:
            continue
        next_non_empty = 0
        if row < bounds.max_row:
            next_non_empty = sum(1 for value in _row_values(worksheet, row + 1, bounds) if value)
        score = len(non_empty) + min(next_non_empty, len(non_empty))
        if score >= HEADER_SCORE_MIN * 2:
            headers.append(
                {
                    "row": row,
                    "labels": non_empty,
                    "score": score,
                    "columns": [
                        {"column": col, "label": _as_text(worksheet.cell(row, col).value)}
                        for col in range(bounds.min_col, bounds.max_col + 1)
                        if _as_text(worksheet.cell(row, col).value)
                    ],
                }
            )
    return headers[:20]


def _infer_field_type(label: str, sample_values: list[Any]) -> str:
    normalized = label.lower().strip()
    if any(hint in normalized for hint in PHOTO_HINTS):
        return "file"
    if any(hint in normalized for hint in SIGNATURE_HINTS):
        return "text"
    if any(hint in normalized for hint in LONG_TEXT_HINTS):
        return "textarea"
    if any(hint in normalized for hint in DATE_HINTS):
        return "date"
    if any(hint in normalized for hint in NUMBER_HINTS):
        return "number"
    non_empty = [value for value in sample_values if value not in (None, "")]
    if non_empty and all(_looks_checkbox(value) for value in non_empty):
        return "checkbox"
    if non_empty and all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in non_empty):
        return "number"
    if non_empty and all(isinstance(value, (date, datetime)) for value in non_empty):
        return "date"
    return "text"


def _looks_checkbox(value) -> bool:
    if isinstance(value, bool):
        return True
    return _as_text(value).lower() in CHECKBOX_VALUES


def _detect_tables(worksheet, bounds: SheetBounds, header_rows: list[dict]) -> list[dict]:
    tables = []
    for header in header_rows:
        row = header["row"]
        columns = []
        for column in header["columns"]:
            col = column["column"]
            label = column["label"]
            sample_values = [worksheet.cell(r, col).value for r in range(row + 1, min(row + 16, bounds.max_row) + 1)]
            columns.append(
                {
                    "source_ref": worksheet.cell(row, col).coordinate,
                    "column": col,
                    "label": label,
                    "key": safe_unique_key(label, {item["key"] for item in columns}),
                    "field_type": _infer_field_type(label, sample_values),
                }
            )
        if len(columns) >= 2:
            data_end = row
            for candidate in range(row + 1, bounds.max_row + 1):
                values = [_as_text(worksheet.cell(candidate, col["column"]).value) for col in columns]
                if not any(values):
                    break
                data_end = candidate
            tables.append(
                {
                    "name": f"Tabelle ab Zeile {row}",
                    "header_row": row,
                    "data_start_row": row + 1,
                    "data_end_row": data_end,
                    "columns": columns,
                    "confidence": min(0.95, 0.35 + 0.1 * len(columns)),
                }
            )
    return _dedupe_overlapping_tables(tables)


def _dedupe_overlapping_tables(tables: list[dict]) -> list[dict]:
    result = []
    occupied_rows = set()
    for table in sorted(tables, key=lambda item: (-item["confidence"], item["header_row"])):
        table_rows = set(range(table["header_row"], table["data_end_row"] + 1))
        if occupied_rows.intersection(table_rows):
            continue
        occupied_rows.update(table_rows)
        result.append(table)
    return sorted(result, key=lambda item: item["header_row"])[:12]


def _detect_checkbox_like_cells(worksheet, bounds: SheetBounds) -> list[dict]:
    cells = []
    for row in range(bounds.min_row, bounds.max_row + 1):
        for col in range(bounds.min_col, bounds.max_col + 1):
            cell = worksheet.cell(row, col)
            if _looks_checkbox(cell.value):
                cells.append({"cell": cell.coordinate, "value": _as_text(cell.value)})
    return cells


def _detect_date_like_cells(worksheet, bounds: SheetBounds) -> list[dict]:
    cells = []
    for row in range(bounds.min_row, bounds.max_row + 1):
        for col in range(bounds.min_col, bounds.max_col + 1):
            cell = worksheet.cell(row, col)
            value = cell.value
            number_format = str(getattr(cell, "number_format", "") or "").lower()
            if isinstance(value, (date, datetime)) or any(hint in number_format for hint in ["yy", "dd", "mm"]):
                cells.append({"cell": cell.coordinate, "value": _as_text(_cell_value(cell))})
    return cells


def _is_inside_table(row: int, col: int, tables: list[dict]) -> bool:
    for table in tables:
        if table["header_row"] <= row <= table["data_end_row"]:
            table_cols = {column["column"] for column in table.get("columns", [])}
            if col in table_cols:
                return True
    return False


def _detect_empty_input_cells(worksheet, bounds: SheetBounds, tables: list[dict]) -> list[dict]:
    inputs = []
    for row in range(bounds.min_row, bounds.max_row + 1):
        for col in range(bounds.min_col, bounds.max_col + 1):
            if _is_inside_table(row, col, tables):
                continue
            cell = worksheet.cell(row, col)
            if cell.value not in (None, ""):
                continue
            left = _as_text(worksheet.cell(row, col - 1).value) if col > bounds.min_col else ""
            above = _as_text(worksheet.cell(row - 1, col).value) if row > bounds.min_row else ""
            label = left or above
            if label and any(ch.isalpha() for ch in label):
                inputs.append({"cell": cell.coordinate, "label": label[:120], "source": "left" if left else "above"})
    return inputs


def _detected_fields_from_inputs(inputs: list[dict], date_cells: list[dict], checkbox_cells: list[dict]) -> list[dict]:
    fields = []
    used = set()
    date_refs = {item["cell"] for item in date_cells}
    checkbox_refs = {item["cell"] for item in checkbox_cells}
    for item in inputs:
        label = item["label"]
        key = safe_unique_key(label, used)
        used.add(key)
        field_type = _infer_field_type(label, [])
        if item["cell"] in date_refs:
            field_type = "date"
        if item["cell"] in checkbox_refs:
            field_type = "checkbox"
        fields.append({"label": label, "key": key, "field_type": field_type, "source_ref": item["cell"]})
    return fields


def safe_unique_key(label: str, used: set[str] | None = None) -> str:
    used = used or set()
    raw = slugify(label or "feld") or "feld"
    raw = re.sub(r"[^a-z0-9_-]+", "-", raw.lower()).strip("-")[:60] or "feld"
    key = raw
    counter = 2
    while key in used:
        suffix = f"-{counter}"
        key = f"{raw[: 80 - len(suffix)]}{suffix}"
        counter += 1
    return key[:80]


def build_default_mapping(analysis: dict, *, mode: str = "one_form_per_sheet") -> dict:
    selected = [sheet.get("name") for sheet in analysis.get("sheets", []) if sheet.get("row_count", 0) > 0]
    return {
        "version": 1,
        "mode": mode,
        "form_title": os.path.splitext(analysis.get("source_filename", "Excel Import"))[0] or "Excel Import",
        "selected_sheets": selected,
        "sheets": [_mapping_for_sheet(sheet) for sheet in analysis.get("sheets", [])],
    }


def _mapping_for_sheet(sheet: dict) -> dict:
    title = sheet.get("title_cells", [{}])[0].get("value") if sheet.get("title_cells") else sheet.get("name")
    fields = sheet.get("detected_fields") or []
    tables = []
    for table in sheet.get("tables", [])[:3]:
        tables.append(
            {
                "title": table.get("name") or "Tabelle",
                "key": safe_unique_key(table.get("name") or "tabelle", set()),
                "source_ref": f"row:{table.get('header_row')}",
                "min_rows": 0,
                "max_rows": 100,
                "columns": table.get("columns", []),
            }
        )
    return {
        "name": sheet.get("name"),
        "title": title or sheet.get("name"),
        "selected": sheet.get("row_count", 0) > 0,
        "fields": fields,
        "tables": tables,
    }


def mapping_to_pretty_json(mapping: dict) -> str:
    return json.dumps(mapping or {}, ensure_ascii=False, indent=2, sort_keys=True)


def parse_mapping_json(raw: str) -> dict:
    try:
        mapping = json.loads(raw or "{}")
    except json.JSONDecodeError as exc:
        raise ValidationError(f"Mapping ist kein gueltiges JSON: {exc}") from exc
    validate_mapping(mapping)
    return mapping


def validate_mapping(mapping: dict) -> None:
    if not isinstance(mapping, dict):
        raise ValidationError("Mapping muss ein JSON-Objekt sein.")
    if mapping.get("mode") not in {"one_form_per_sheet", "all_sheets_one_form"}:
        raise ValidationError("Bitte einen gueltigen Importmodus waehlen.")
    if not isinstance(mapping.get("sheets", []), list):
        raise ValidationError("Mapping muss eine sheets-Liste enthalten.")
    selected = set(mapping.get("selected_sheets") or [])
    for sheet in mapping.get("sheets", []):
        if not isinstance(sheet, dict):
            raise ValidationError("Jedes Sheet-Mapping muss ein Objekt sein.")
        if sheet.get("name") in selected and not sheet.get("title"):
            raise ValidationError("Ausgewaehlte Blaetter brauchen einen Titel.")
        for field in sheet.get("fields", []):
            if field.get("field_type") not in {"text", "textarea", "checkbox", "date", "number", "select", "file"}:
                raise ValidationError("Ein Feldmapping enthaelt einen nicht erlaubten Feldtyp.")
        for table in sheet.get("tables", []):
            for column in table.get("columns", []):
                if column.get("field_type") not in {"text", "textarea", "checkbox", "date", "number", "select", "file"}:
                    raise ValidationError("Eine Tabellenspalte enthaelt einen nicht erlaubten Feldtyp.")


def save_mapping(*, job: ImportJob, mapping: dict) -> ImportJob:
    validate_mapping(mapping)
    job.mapping = mapping
    job.status = ImportJob.ImportStatus.MAPPED
    job.error_message = ""
    job.save(update_fields=["mapping", "status", "error_message", "updated_at"])
    return job


def generate_draft_forms_from_mapping(*, job: ImportJob, user) -> list[Form]:
    mapping = job.mapping or build_default_mapping(job.analysis_result or {})
    validate_mapping(mapping)
    selected_names = set(mapping.get("selected_sheets") or [])
    sheets = [sheet for sheet in mapping.get("sheets", []) if sheet.get("name") in selected_names]
    if not sheets:
        raise ValidationError("Bitte mindestens ein Excel-Blatt fuer die Generierung auswaehlen.")
    with transaction.atomic():
        generated = (
            [_generate_combined_form(job=job, mapping=mapping, sheets=sheets, user=user)]
            if mapping.get("mode") == "all_sheets_one_form"
            else [_generate_form_for_sheet(job=job, mapping=mapping, sheet=sheet, user=user) for sheet in sheets]
        )
        job.generated_form_ids = [str(form.pk) for form in generated]
        job.status = ImportJob.ImportStatus.GENERATED
        job.error_message = ""
        job.save(update_fields=["generated_form_ids", "status", "error_message", "updated_at"])
    return generated


def _next_form_version(key: str) -> int:
    latest = Form.objects.filter(key=key).order_by("-version").first()
    return int(latest.version + 1) if latest else 1


def _unique_form_key(base: str) -> str:
    key = safe_unique_key(base, set())
    if not Form.objects.filter(key=key).exists():
        return key
    counter = 2
    while True:
        candidate = f"{key[:72]}-{counter}"
        if not Form.objects.filter(key=candidate).exists():
            return candidate
        counter += 1


def _create_form(*, title: str, user, source: dict) -> Form:
    key = _unique_form_key(title)
    form = Form.objects.create(
        key=key,
        version=_next_form_version(key),
        title=title[:255],
        description="Aus Excel-Import erzeugter Formularentwurf.",
        status=Form.PublicationStatus.DRAFT,
        review_required=True,
        is_archivable=True,
        schema={},
        created_by=user,
        updated_by=user,
    )
    form.schema = {"excel_import_source": source}
    form.save(update_fields=["schema", "updated_at"])
    return form


def _generate_form_for_sheet(*, job: ImportJob, mapping: dict, sheet: dict, user) -> Form:
    title = sheet.get("title") or sheet.get("name") or mapping.get("form_title") or "Excel Import"
    form = _create_form(title=title, user=user, source={"job_id": str(job.pk), "sheet": sheet.get("name")})
    section = FormSection.objects.create(
        form=form,
        title=title[:255],
        description="Aus Excel-Blatt erzeugter Abschnitt.",
        position=1,
        created_by=user,
        updated_by=user,
    )
    _create_fields_and_tables(form=form, section=section, sheet=sheet, job=job, user=user)
    form.sync_schema()
    _audit_generated_form(job=job, form=form, user=user)
    return form


def _generate_combined_form(*, job: ImportJob, mapping: dict, sheets: list[dict], user) -> Form:
    title = mapping.get("form_title") or job.original_filename or "Excel Import"
    form = _create_form(title=title, user=user, source={"job_id": str(job.pk), "mode": "all_sheets_one_form"})
    for index, sheet in enumerate(sheets, start=1):
        section = FormSection.objects.create(
            form=form,
            title=(sheet.get("title") or sheet.get("name") or f"Blatt {index}")[:255],
            description="Aus Excel-Blatt erzeugter Abschnitt.",
            position=index,
            created_by=user,
            updated_by=user,
        )
        _create_fields_and_tables(form=form, section=section, sheet=sheet, job=job, user=user)
    form.sync_schema()
    _audit_generated_form(job=job, form=form, user=user)
    return form


def _field_type_for_model(field_type: str) -> str:
    return {
        "checkbox": Field.FieldType.BOOLEAN,
        "number": Field.FieldType.DECIMAL,
        "table": Field.FieldType.TEXT,
    }.get(field_type, field_type)


def _column_type_for_model(field_type: str) -> str:
    return {
        "checkbox": RepeatableGroupColumn.ColumnType.BOOLEAN,
        "number": RepeatableGroupColumn.ColumnType.DECIMAL,
    }.get(field_type, field_type)


def _next_field_position(form: Form) -> int:
    last_position = (
        Field.objects.filter(form=form).order_by("-position").values_list("position", flat=True).first()
    )
    return int(last_position or 0) + 1


def _create_fields_and_tables(*, form: Form, section: FormSection, sheet: dict, job: ImportJob, user) -> None:
    used_field_keys = set(Field.objects.filter(form=form).values_list("key", flat=True))
    position = _next_field_position(form)
    for field in sheet.get("fields", [])[:80]:
        label = field.get("label") or field.get("key") or "Feld"
        key = safe_unique_key(field.get("key") or label, used_field_keys)
        used_field_keys.add(key)
        Field.objects.create(
            form=form,
            section=section,
            key=key,
            label=label[:255],
            field_type=_field_type_for_model(field.get("field_type", "text")),
            position=position,
            required=bool(field.get("required", False)),
            ui_config={"excel_import": {"job_id": str(job.pk), "sheet": sheet.get("name"), "source_ref": field.get("source_ref")}},
            created_by=user,
            updated_by=user,
        )
        position += 1
    for table in sheet.get("tables", [])[:10]:
        _create_repeatable_group(form=form, section=section, table=table, job=job, sheet=sheet, user=user)


def _create_repeatable_group(*, form: Form, section: FormSection, table: dict, job: ImportJob, sheet: dict, user) -> RepeatableGroup:
    existing_group_keys = set(RepeatableGroup.objects.filter(form=form).values_list("key", flat=True))
    group_key = safe_unique_key(table.get("key") or table.get("title") or "tabelle", existing_group_keys)
    group = RepeatableGroup.objects.create(
        form=form,
        section=section,
        key=group_key,
        title=(table.get("title") or "Tabelle")[:255],
        description="Aus Excel-Tabelle erzeugte wiederholbare Eingabe.",
        position=int(table.get("position") or 1),
        min_rows=int(table.get("min_rows") or 0),
        max_rows=min(int(table.get("max_rows") or 100), 200),
        ui_config={"excel_import": {"job_id": str(job.pk), "sheet": sheet.get("name"), "source_ref": table.get("source_ref")}},
        created_by=user,
        updated_by=user,
    )
    used_columns = set()
    nicht_ok_key = None
    conditional_required_keys = []
    for index, column in enumerate(table.get("columns", []), start=1):
        label = column.get("label") or f"Spalte {index}"
        key = safe_unique_key(column.get("key") or label, used_columns)
        used_columns.add(key)
        field_type = column.get("field_type", "text")
        rules = dict(column.get("validation_rules") or {})
        label_lower = label.lower()
        if "nicht" in label_lower and "ok" in label_lower:
            nicht_ok_key = key
            field_type = "checkbox"
        if any(hint in label_lower for hint in ["bemerkung", "mangel", "maßnahme", "massnahme"]):
            conditional_required_keys.append(key)
        RepeatableGroupColumn.objects.create(
            group=group,
            key=key,
            label=label[:255],
            column_type=_column_type_for_model(field_type),
            position=index,
            required=bool(column.get("required", False)),
            validation_rules=rules,
            ui_config={"excel_import": {"job_id": str(job.pk), "sheet": sheet.get("name"), "source_ref": column.get("source_ref")}},
            created_by=user,
            updated_by=user,
        )
    if nicht_ok_key and conditional_required_keys:
        for column in group.columns.filter(key__in=conditional_required_keys):
            rules = dict(column.validation_rules or {})
            rules["required_if"] = {"column": nicht_ok_key, "equals": True}
            column.validation_rules = rules
            column.save(update_fields=["validation_rules", "updated_at"])
    return group


def _audit_generated_form(*, job: ImportJob, form: Form, user) -> None:
    AuditLog.objects.create(
        actor=user,
        event_type=AuditLog.EventType.CREATED,
        target_model="Form",
        target_id=form.pk,
        form=form,
        message="Formularentwurf aus Excel-Import erzeugt.",
        metadata={"import_job_id": str(job.pk), "source_filename": job.original_filename},
    )


def build_hygiene_demo_workbook_bytes() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise ValidationError("openpyxl ist nicht installiert.") from exc
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in ["Kueche", "Sanitaer", "Gemeinschaft"]:
        ws = workbook.create_sheet(sheet_name)
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Hygiene Kontrolle {sheet_name}"
        headers = ["Bereich", "Kontrollpunkt", "OK", "Nicht OK", "Bemerkung", "Maßnahme", "Verantwortlich", "Frist"]
        for col, label in enumerate(headers, start=1):
            ws.cell(3, col).value = label
        for row in range(4, 9):
            ws.cell(row, 1).value = sheet_name
            ws.cell(row, 2).value = f"Kontrollpunkt {row - 3}"
            ws.cell(row, 3).value = "☐"
            ws.cell(row, 4).value = "☐"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
