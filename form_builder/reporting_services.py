from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Any

from django.db.models import Count, Q
from django.utils import timezone

from .action_item_models import ActionItem
from .models import AuditLog, Form, FormEntry, OutboxItem
from .permissions import entry_scope_q


@dataclass(frozen=True)
class ReportingPeriod:
    start: datetime
    end: datetime


def _local_day_bounds(value: date) -> tuple[datetime, datetime]:
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.combine(value, time.min), tz)
    end = timezone.make_aware(datetime.combine(value, time.max), tz)
    return start, end


def default_month_period(today: date | None = None) -> ReportingPeriod:
    today = today or timezone.localdate()
    start_date = today.replace(day=1)
    if start_date.month == 12:
        next_month = start_date.replace(year=start_date.year + 1, month=1)
    else:
        next_month = start_date.replace(month=start_date.month + 1)
    start, _ = _local_day_bounds(start_date)
    end, _ = _local_day_bounds(next_month)
    return ReportingPeriod(start=start, end=end)


def parse_date_period(
    start_text: str | None = None, end_text: str | None = None
) -> ReportingPeriod:
    if not start_text and not end_text:
        return default_month_period()
    start_date = (
        date.fromisoformat(start_text) if start_text else timezone.localdate().replace(day=1)
    )
    end_date = date.fromisoformat(end_text) if end_text else timezone.localdate()
    start, _ = _local_day_bounds(start_date)
    _, end = _local_day_bounds(end_date)
    return ReportingPeriod(start=start, end=end)


def scoped_entries(user, *, form: Form | None = None, period: ReportingPeriod | None = None):
    queryset = FormEntry.objects.select_related(
        "form", "bewohner", "created_by", "updated_by"
    ).exclude(status=FormEntry.EntryStatus.DELETED)
    if user is not None:
        queryset = queryset.filter(entry_scope_q(user))
    if form is not None:
        queryset = queryset.filter(form=form)
    if period is not None:
        queryset = queryset.filter(created_at__gte=period.start, created_at__lte=period.end)
    return queryset.order_by("form__title", "created_at")


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Ja" if value else "Nein"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, list):
        return ", ".join(_display_value(item) for item in value)
    if isinstance(value, dict):
        for key in ("filename", "value", "label", "title"):
            if value.get(key):
                return str(value[key])
        return ""
    return str(value)


def _field_definitions_for_entries(entries: list[FormEntry]) -> list[dict]:
    by_key: dict[str, dict] = {}
    for entry in entries:
        for field in (entry.form_snapshot or {}).get("fields", []):
            key = field.get("key")
            if key and key not in by_key:
                by_key[key] = field
    return list(by_key.values())


def _repeatable_groups_for_entries(entries: list[FormEntry]) -> list[dict]:
    groups: dict[str, dict] = {}
    for entry in entries:
        for group in (entry.form_snapshot or {}).get("repeatable_groups", []):
            key = group.get("key")
            if key and key not in groups:
                groups[key] = group
    return list(groups.values())


def export_entries_to_xlsx(*, user, entries, filename_prefix: str = "formulare") -> bytes:
    from openpyxl import Workbook

    entry_list = list(entries)
    wb = Workbook()
    ws = wb.active
    ws.title = "Eintraege"
    fields = _field_definitions_for_entries(entry_list)
    headers = [
        "Entry ID",
        "Public ID",
        "Formular",
        "Form Key",
        "Status",
        "Bewohner",
        "Erstellt von",
        "Aktualisiert von",
        "Eingereicht am",
        "Erstellt am",
    ] + [field.get("label") or field.get("key") for field in fields]
    ws.append(headers)
    for entry in entry_list:
        data = entry.data or {}
        ws.append(
            [
                str(entry.pk),
                str(entry.public_id),
                entry.form.title,
                entry.form.key,
                entry.get_status_display(),
                str(entry.bewohner),
                getattr(entry.created_by, "username", "") if entry.created_by_id else "",
                getattr(entry.updated_by, "username", "") if entry.updated_by_id else "",
                timezone.localtime(entry.submitted_at).isoformat() if entry.submitted_at else "",
                timezone.localtime(entry.created_at).isoformat() if entry.created_at else "",
            ]
            + [_display_value(data.get(field.get("key"))) for field in fields]
        )

    for group in _repeatable_groups_for_entries(entry_list):
        group_key = group.get("key")
        if not group_key:
            continue
        title = (group.get("title") or group_key)[:28]
        sheet_name = title or "Tabelle"
        suffix = 1
        while sheet_name in wb.sheetnames:
            suffix += 1
            sheet_name = f"{title[:25]} {suffix}"
        table_ws = wb.create_sheet(sheet_name)
        columns = group.get("columns") or []
        table_ws.append(
            ["Entry ID", "Formular", "Bewohner", "Zeile"]
            + [column.get("label") or column.get("key") for column in columns]
        )
        for entry in entry_list:
            rows = (entry.data or {}).get(group_key) or []
            if not isinstance(rows, list):
                continue
            for index, row in enumerate(rows, start=1):
                if not isinstance(row, dict):
                    continue
                table_ws.append(
                    [str(entry.pk), entry.form.title, str(entry.bewohner), index]
                    + [_display_value(row.get(column.get("key"))) for column in columns]
                )

    for worksheet in wb.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
        worksheet.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    (
        AuditLog.objects.create(
            actor=user,
            event_type=AuditLog.EventType.DOWNLOAD,
            target_model="FormEntry",
            target_id=entry_list[0].pk if entry_list else FormEntry.objects.none().model().pk,
            form=entry_list[0].form if entry_list else None,
            form_entry=entry_list[0] if entry_list else None,
            message="Formulareintraege wurden als Excel exportiert.",
            metadata={"count": len(entry_list), "filename_prefix": filename_prefix},
        )
        if entry_list
        else None
    )
    return buffer.getvalue()


def get_operational_dashboard_data(user) -> dict:
    now = timezone.now()
    today = timezone.localdate()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    today_start, today_end = _local_day_bounds(today)
    week_start_dt, _ = _local_day_bounds(week_start)
    month_start_dt, _ = _local_day_bounds(month_start)

    entries = scoped_entries(user)
    action_items = ActionItem.objects.select_related(
        "source_entry", "source_entry__form", "assigned_to"
    )
    if user is not None:
        action_items = action_items.filter(source_entry__in=scoped_entries(user).values("pk"))

    pending_reviews = entries.filter(status=FormEntry.EntryStatus.IN_REVIEW).order_by(
        "submitted_at"
    )[:10]
    ready_to_send = entries.filter(status=FormEntry.EntryStatus.APPROVED).order_by("updated_at")[
        :10
    ]
    failed_outbox = OutboxItem.objects.select_related(
        "form_entry", "form", "bewohner", "recipient"
    ).filter(status=OutboxItem.DeliveryStatus.FAILED)
    if user is not None:
        failed_outbox = failed_outbox.filter(form_entry__in=scoped_entries(user).values("pk"))

    open_action_statuses = [
        ActionItem.Status.OPEN,
        ActionItem.Status.IN_PROGRESS,
    ]
    overdue_action_items = action_items.filter(
        status__in=open_action_statuses, due_at__lt=now
    ).order_by("due_at")[:10]
    recent_audit = AuditLog.objects.select_related("actor", "form", "form_entry").order_by(
        "-occurred_at"
    )
    if user is not None:
        recent_audit = recent_audit.filter(
            Q(form_entry__isnull=True) | Q(form_entry__in=scoped_entries(user).values("pk"))
        )

    return {
        "counts": {
            "pending_reviews": entries.filter(status=FormEntry.EntryStatus.IN_REVIEW).count(),
            "ready_to_send": entries.filter(status=FormEntry.EntryStatus.APPROVED).count(),
            "failed_outbox": failed_outbox.count(),
            "overdue_action_items": action_items.filter(
                status__in=open_action_statuses, due_at__lt=now
            ).count(),
            "submitted_today": entries.filter(
                submitted_at__gte=today_start, submitted_at__lte=today_end
            ).count(),
            "submitted_week": entries.filter(
                submitted_at__gte=week_start_dt, submitted_at__lte=now
            ).count(),
            "submitted_month": entries.filter(
                submitted_at__gte=month_start_dt, submitted_at__lte=now
            ).count(),
        },
        "pending_reviews": pending_reviews,
        "ready_to_send": ready_to_send,
        "failed_outbox": failed_outbox.order_by("failed_at", "updated_at")[:10],
        "overdue_action_items": overdue_action_items,
        "recent_audit_events": recent_audit[:10],
    }


def monthly_report_context(*, user, form: Form | None, period: ReportingPeriod) -> dict:
    entries = scoped_entries(user, form=form, period=period)
    action_items = ActionItem.objects.select_related("source_entry", "source_entry__form")
    if form is not None:
        action_items = action_items.filter(source_entry__form=form)
    if user is not None:
        action_items = action_items.filter(source_entry__in=scoped_entries(user).values("pk"))
    action_items = action_items.filter(created_at__gte=period.start, created_at__lte=period.end)
    open_statuses = [ActionItem.Status.OPEN, ActionItem.Status.IN_PROGRESS]
    return {
        "form": form,
        "period": period,
        "entries": entries,
        "total_submissions": entries.count(),
        "open_issues": action_items.filter(status__in=open_statuses).count(),
        "completed_issues": action_items.filter(
            status__in=[ActionItem.Status.DONE, ActionItem.Status.VERIFIED]
        ).count(),
        "overdue_issues": action_items.filter(
            status__in=open_statuses, due_at__lt=timezone.now()
        ).count(),
        "critical_items": action_items.filter(priority=ActionItem.Priority.HIGH).order_by("due_at")[
            :25
        ],
        "per_form_summary": list(
            entries.values("form__title").annotate(total=Count("id")).order_by("form__title")
        ),
    }


def render_monthly_report_pdf(
    *, user, form: Form | None, period: ReportingPeriod | None = None
) -> bytes:
    period = period or default_month_period()
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    context = monthly_report_context(user=user, form=form, period=period)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=16 * mm, rightMargin=16 * mm)
    styles = getSampleStyleSheet()
    title = "Monatsbericht"
    if form is not None:
        title += f" - {form.title}"
    story = [Paragraph(title, styles["Title"]), Spacer(1, 5 * mm)]
    story.append(
        Paragraph(
            f"Zeitraum: {timezone.localtime(period.start).date()} bis {timezone.localtime(period.end).date()}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 5 * mm))
    metrics = [
        ["Kennzahl", "Wert"],
        ["Einreichungen", context["total_submissions"]],
        ["Offene Massnahmen", context["open_issues"]],
        ["Erledigte Massnahmen", context["completed_issues"]],
        ["Ueberfaellige Massnahmen", context["overdue_issues"]],
    ]
    table = Table(metrics, colWidths=[90 * mm, 70 * mm])
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("Kritische Massnahmen", styles["Heading2"]))
    critical = [["Titel", "Status", "Faellig", "Formular"]]
    for item in context["critical_items"]:
        critical.append(
            [
                item.title,
                item.get_status_display(),
                timezone.localtime(item.due_at).date().isoformat() if item.due_at else "-",
                item.source_entry.form.title if item.source_entry_id else "-",
            ]
        )
    if len(critical) == 1:
        critical.append(["Keine kritischen Massnahmen", "", "", ""])
    critical_table = Table(critical, colWidths=[70 * mm, 35 * mm, 35 * mm, 45 * mm])
    critical_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ]
        )
    )
    story.append(critical_table)
    doc.build(story)
    (
        AuditLog.objects.create(
            actor=user,
            event_type=AuditLog.EventType.DOWNLOAD,
            target_model="Form" if form else "FormEntry",
            target_id=(
                form.pk
                if form
                else (
                    context["entries"].first().pk
                    if context["entries"].exists()
                    else Form.objects.first().pk
                )
            ),
            form=form,
            message="Monatsbericht wurde als PDF erzeugt.",
            metadata={"start": period.start.isoformat(), "end": period.end.isoformat()},
        )
        if (form or context["entries"].exists() or Form.objects.exists())
        else None
    )
    return buffer.getvalue()
